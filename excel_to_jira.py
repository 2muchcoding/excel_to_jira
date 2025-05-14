import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
import warnings
import os
from openpyxl import load_workbook
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# === Custom fields like epic name and link ===
def getCustomFieldIds():
    epic_link_field = None
    epic_name_field = None
    risk_field = None

    field_response = requests.get(
        f"{jira_url}/rest/api/2/field",
        headers=headers,
        auth=auth
    )
    fields = field_response.json()

    for field in fields:
        if field.get('custom'):
            if field['name'] == "Epic Link":
                epic_link_field = field['id']
            elif field['name'] == "Epic Name":
                epic_name_field = field['id']
            elif field['name'] == "Risk level":
                risk_field = field['id']
    if not all([epic_link_field, epic_name_field, risk_field]):
        print("Could not find all required custom fields.")
        exit()

    return epic_link_field, epic_name_field, risk_field

# === Check if content in the cell exists ===
def checkContentExists(content):
    if pd.isna(content) or str(content).lower() == "nan":
            return "n/a"
    return str(content)

# === Create Epic and Tasks ===
def createEpic (wb, gate_name, project_key, epic_link_field, risk_field):
    print("Creating Epic and Tasks...")
    
    sheet = wb[gate_name]
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]
    df = df.iloc[1:]
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    
    # === Create Epic ===
    title = str(df.iloc[2, 3]).strip()  
    summary = f"{gate_name} - {title}"

    epic_data = {
        "fields": {
            "project": { "key": project_key },
            "summary": summary,
            "issuetype": { "name": "Epic" },
            epic_name_field: summary,
            "duedate": "2025-12-31"
        }
    }

    epic_response = requests.post(
        f"{jira_url}/rest/api/2/issue",
        headers=headers,
        auth=auth,
        json=epic_data
    )

    if epic_response.status_code not in [200, 201]:
        print(f"Failed to create epic: {epic_response.status_code} - {epic_response.text}")
        exit()
    else:
        print(f"Successfully created epic: {summary}")

    epic_key = epic_response.json().get("key")

    # === Create Tasks ===
    for index in range(3, len(df)):
        gate_num = str(df.iloc[index, 1]).strip()
        category = str(df.iloc[index, 2]).strip()

        task = checkContentExists(df.iloc[index, 3].strip())

        template_and_guidlines_text = str(df.iloc[index, 4].strip())
        cell_template = sheet.cell(row=index + 2, column=5)
        if cell_template.hyperlink:
            url = cell_template.hyperlink.target
            if url.startswith("../"):
                url = "https://jira.indiesemi.com:8443/" + url.lstrip("../")
            template_and_guidlines = f"[{template_and_guidlines_text}|{url}]"
        else:
            template_and_guidlines = checkContentExists(template_and_guidlines_text)

        risk = str(df.iloc[index, 37]).strip()
        if pd.isna(risk) or str(risk).lower() == "nan":
            risk = None
        risk_map = {"Low/No Risk": "Low",
                    "Low": "Low",
                    "Medium": "Medium",
                    "High": "High"}
        risk = risk_map.get(risk, None)

        results = checkContentExists(df.iloc[index, 35].strip())

        links_to_evidence_text = str(df.iloc[index, 36].strip())
        cell_links = sheet.cell(row=index + 2, column=37)
        if cell_links.hyperlink:
            url = cell_links.hyperlink.target
            if url.startswith("../"):
                url = "https://jira.indiesemi.com:8443/" + url.lstrip("../")
            links_to_evidence = f"[{links_to_evidence_text}|{url}]"
        else:
            links_to_evidence = checkContentExists(links_to_evidence_text)

        task_data = {
            "fields": {
                "project": {
                    "key": project_key
                },
                "summary": gate_num + " " + category,
                "description": "*Task:* " + task + "\n\n" + 
                               "*Template+Guidelines:* " + template_and_guidlines + "\n\n" +
                               "*Results:* " + results + "\n\n" +
                               "*Links to Evidence:* " + links_to_evidence,
                "issuetype": {
                    "name": "Task"
                },
                epic_link_field: epic_key,  
                "duedate": "2025-12-31",
            }
        }
        if risk:
            task_data["fields"][risk_field] = {"value": risk}

        task_response = requests.post(
            f"{jira_url}/rest/api/2/issue",
            headers=headers,
            auth=auth,
            json=task_data
        )

        if task_response.status_code not in [200, 201]:
            print(f"Failed to create task for {gate_num}: {task_response.status_code} - {task_response.text}")
        else:
            print(f"Successfully created task: {gate_num} {category}")

def checkCredentials ():
    # === Check Credentials ===
    test_response = requests.get(
        f"{jira_url}/rest/api/2/myself",
        headers=headers,
        auth=auth,
    )
    if test_response.status_code != 200:
        return False
    return True


if __name__ == "__main__":
        # === Jira Configuration ===
    jira_url = "https://jira.indiesemi.com:8443"

    # === User Input ===
    pat_choice = input("Use your JIRA access token (PAT) instead of password? (y/n): ").strip().lower()
    
    if pat_choice not in ["y", "n"]:
        print("Invalid choice. Please enter 'y' or 'n'.")
        exit()
    
    username = input("Enter your Jira username: ").strip()
    if not username:
        print("Username cannot be empty.")
        exit()

    if pat_choice == "y":
        token = input("Enter your Jira access token (PAT): ").strip()
        if not token:
            print("Token cannot be empty.")
            exit()
        auth = None
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
    elif pat_choice == "n":
        password = input("Enter your Jira password: ").strip()
        if not password:
            print("Password cannot be empty.")
            exit()
        auth = HTTPBasicAuth(username, password)
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
    if not checkCredentials():
        print("Invalid credentials. Please check your username and password/token.")
        exit()

    # === Get Projects ===
    if input("Show projects? (y/n): ").strip().lower() == "y":
        project_response = requests.get(
            f"{jira_url}/rest/api/2/project",
            headers=headers,
            auth=auth
        )
        if project_response.status_code == 200:
            projects = project_response.json()
            print("Available projects:")
            for project in projects:
                print(f"- {project['key']}: {project['name']}")
        else:
            print("Failed to retrieve projects:", project_response.status_code, project_response.text)
            exit()

    project_response = requests.get(
        f"{jira_url}/rest/api/2/project",
        headers=headers,
        auth=auth
    )
    project_key = input("Enter the project key: ").strip().upper()
    if project_key not in [project['key'] for project in project_response.json()]:
        print(f"Project '{project_key}' not found.")
        exit()
     
    file_path = input("Enter the path to the Excel file: ").strip()
    if not os.path.isfile(file_path):
        print(f"Error: The file '{file_path}' does not exist. Please check the path and try again.")
        exit()

    epic_link_field, epic_name_field, risk_field = getCustomFieldIds()

    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    gate_sheets = [sheet_name for sheet_name in wb.sheetnames if sheet_name.startswith("G")]

    if input("Show all available gates? (y/n): ").strip().lower() == "y":
        print("All available gates:")
        for gate in gate_sheets:
            if gate.startswith("G"):
                print(f"- {gate}")

    gate_name = input("Enter the gate ticket you want to add in JIRA: ").strip()
    if gate_name not in wb.sheetnames:
        print(f"Gate '{gate_name}' not found in the Excel file.")
        exit()
    
    createEpic(wb, gate_name, project_key, epic_link_field, risk_field)