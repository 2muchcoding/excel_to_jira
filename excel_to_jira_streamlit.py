
import streamlit as st
import pandas as pd
import requests
from requests.auth import HTTPBasicAuth
import warnings
from openpyxl import load_workbook
import io

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

st.title("Excel to Jira Uploader")

# User inputs
jira_url = "https://jira.indiesemi.com:8443"

auth_method = st.radio("Authentication method:", ["Username + Password", "Access Token (PAT)"])
username = st.text_input("Enter your Jira username:")
auth = None
headers = {}

if auth_method == "Access Token (PAT)":
    token = st.text_input("Enter your Jira access token (PAT):", type="password")
    if token:
        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
else:
    password = st.text_input("Enter your Jira password:", type="password")
    if username and password:
        auth = HTTPBasicAuth(username, password)
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json"
        }

# Check Jira credentials
def checkCredentials():
    test_response = requests.get(f"{jira_url}/rest/api/2/myself", headers=headers, auth=auth)
    return test_response.status_code == 200

def getCustomFieldIds():
    epic_link_field = None
    epic_name_field = None
    risk_field = None

    field_response = requests.get(f"{jira_url}/rest/api/2/field", headers=headers, auth=auth)
    fields = field_response.json()

    for field in fields:
        if field.get('custom'):
            if field['name'] == "Epic Link":
                epic_link_field = field['id']
            elif field['name'] == "Epic Name":
                epic_name_field = field['id']
            elif field['name'] == "Risk level":
                risk_field = field['id']

    return epic_link_field, epic_name_field, risk_field

def checkContentExists(content):
    if pd.isna(content) or str(content).lower() == "nan":
        return "n/a"
    return str(content)

def createEpicAndTasks(file_bytes, gate_name, project_key, epic_link_field, epic_name_field, risk_field):
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=False)
    sheet = wb[gate_name]
    df = pd.DataFrame(sheet.values)
    df.columns = df.iloc[0]
    df = df.iloc[1:]

    title = str(df.iloc[2, 3]).strip()
    summary = f"{gate_name} - {title}"

    epic_data = {
        "fields": {
            "project": { "key": project_key },
            "summary": summary,
            "issuetype": { "name": "Epic" },
            epic_name_field: summary,
            "duedate": "2025-12-31"
        }
    }

    epic_response = requests.post(f"{jira_url}/rest/api/2/issue", headers=headers, auth=auth, json=epic_data)

    if epic_response.status_code not in [200, 201]:
        st.error(f"Failed to create epic: {epic_response.status_code} - {epic_response.text}")
        return

    epic_key = epic_response.json().get("key")
    st.success(f"Successfully created epic: {summary}")

    for index in range(3, len(df)):
        gate_num = str(df.iloc[index, 1]).strip()
        category = str(df.iloc[index, 2]).strip()
        task = checkContentExists(df.iloc[index, 3])
        template_text = str(df.iloc[index, 4])
        risk = df.iloc[index, 37]
        results = checkContentExists(df.iloc[index, 35])
        evidence_text = str(df.iloc[index, 36])

        risk_map = {"Low/No Risk": "Low", "Low": "Low", "Medium": "Medium", "High": "High"}
        risk_value = risk_map.get(str(risk).strip(), None)

        template_and_guidelines = checkContentExists(template_text)
        links_to_evidence = checkContentExists(evidence_text)

        task_data = {
            "fields": {
                "project": { "key": project_key },
                "summary": gate_num + " " + category,
                "description": "*Task:* " + task + "\n\n" +
                               "*Template+Guidelines:* " + template_and_guidelines + "\n\n" +
                               "*Results:* " + results + "\n\n" +
                               "*Links to Evidence:* " + links_to_evidence,
                "issuetype": { "name": "Task" },
                epic_link_field: epic_key,
                "duedate": "2025-12-31"
            }
        }

        if risk_value:
            task_data["fields"][risk_field] = {"value": risk_value}

        task_response = requests.post(f"{jira_url}/rest/api/2/issue", headers=headers, auth=auth, json=task_data)

        if task_response.status_code in [200, 201]:
            st.success(f"Created task: {gate_num} {category}")
        else:
            st.warning(f"Failed to create task for {gate_num}: {task_response.status_code}")

# File upload
file = st.file_uploader("Upload your Excel file", type=["xlsx"])
if file and username and headers and (auth or "Authorization" in headers):
    if checkCredentials():
        project_response = requests.get(f"{jira_url}/rest/api/2/project", headers=headers, auth=auth)
        if project_response.status_code == 200:
            projects = project_response.json()
            project_keys = [proj["key"] for proj in projects]
            project_key = st.selectbox("Select Jira Project", project_keys)

            wb = load_workbook(filename=io.BytesIO(file.read()), read_only=True)
            gate_sheets = [sheet_name for sheet_name in wb.sheetnames if sheet_name.startswith("G")]
            gate_name = st.selectbox("Select Gate Sheet", gate_sheets)

            # Rewind file to pass to creator
            file.seek(0)
            if st.button("Create Epic and Tasks"):
                epic_link_field, epic_name_field, risk_field = getCustomFieldIds()
                if all([epic_link_field, epic_name_field, risk_field]):
                    createEpicAndTasks(file.read(), gate_name, project_key, epic_link_field, epic_name_field, risk_field)
                else:
                    st.error("Could not find all required custom fields in Jira.")
        else:
            st.error("Failed to fetch projects from Jira.")
    else:
        st.error("Invalid credentials. Please check and try again.")
